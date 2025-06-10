import yfinance as yf
import pandas as pd
import os
from openpyxl import load_workbook
import streamlit as st

# Path to the stocks.xlsx file
STOCKS_FILE_PATH = 'stocks.xlsx'  # Change this to the correct path if needed

# Load actual dividend data
ACTUAL_DIVIDEND_FILE = 'actual_stock-dividend.csv'

# Function to fetch data for a given stock ticker
def get_financial_data(ticker):
    # Add .NS suffix for Indian stocks
    yf_ticker = ticker + ".NS"
    stock = yf.Ticker(yf_ticker)
    result = {'Ticker': ticker}
    
    try:
        income_statement = stock.financials
        balance_sheet = stock.balance_sheet
        cash_flow = stock.cashflow
        dividends = stock.dividends
    except Exception as e:
        st.error(f"Error fetching financial data for {ticker}: {e}")
        return None

    try:
        historical_data = stock.history(period="1d")
        latest_close_price = historical_data['Close'].iloc[-1]
    except Exception as e:
        latest_close_price = "N/A"

    # Basic financial metrics
    try:
        result['Net Income'] = income_statement.loc['Net Income'].iloc[0] if 'Net Income' in income_statement.index else "N/A"
    except Exception:
        result['Net Income'] = "N/A"
    
    try:
        result['Operating Income'] = income_statement.loc['Operating Income'].iloc[0] if 'Operating Income' in income_statement.index else \
                                   income_statement.loc['EBIT'].iloc[0] if 'EBIT' in income_statement.index else "N/A"
    except Exception:
        result['Operating Income'] = "N/A"
    
    try:
        eps = income_statement.loc['Earnings Before Interest and Taxes'].iloc[0] / stock.info['sharesOutstanding']
    except (KeyError, AttributeError):
        eps = "N/A"
    result['EPS'] = eps
    
    try:
        result['Revenue Growth'] = income_statement.loc['Total Revenue'].pct_change().iloc[-1] if 'Total Revenue' in income_statement.index else "N/A"
    except Exception:
        result['Revenue Growth'] = "N/A"
    
    try:
        result['Retained Earnings'] = balance_sheet.loc['Retained Earnings'].iloc[0] if 'Retained Earnings' in balance_sheet.index else "N/A"
    except Exception:
        result['Retained Earnings'] = "N/A"
    
    try:
        result['Cash Reserves'] = balance_sheet.loc['Cash'].iloc[0] if 'Cash' in balance_sheet.index else "N/A"
    except Exception:
        result['Cash Reserves'] = "N/A"
    
    try:
        result['Debt-to-Equity Ratio'] = balance_sheet.loc['Total Debt'].iloc[0] / balance_sheet.loc['Stockholders Equity'].iloc[0] if 'Total Debt' in balance_sheet.index and 'Stockholders Equity' in balance_sheet.index else "N/A"
    except Exception:
        result['Debt-to-Equity Ratio'] = "N/A"
    
    try:
        result['Working Capital'] = balance_sheet.loc['Total Assets'].iloc[0] - balance_sheet.loc['Total Liabilities Net Minority Interest'].iloc[0] if 'Total Assets' in balance_sheet.index and 'Total Liabilities Net Minority Interest' in balance_sheet.index else "N/A"
    except Exception:
        result['Working Capital'] = "N/A"
    
    try:
        result['Dividend Payout Ratio'] = stock.info.get('dividendYield', "N/A")
    except Exception:
        result['Dividend Payout Ratio'] = "N/A"
    result['Dividend Yield'] = result['Dividend Payout Ratio']
    
    try:
        result['Free Cash Flow'] = cash_flow.loc['Free Cash Flow'].iloc[0] if 'Free Cash Flow' in cash_flow.index else "N/A"
    except Exception:
        result['Free Cash Flow'] = "N/A"
    
    if not isinstance(dividends, pd.Series) or dividends.empty:
        result['Dividend Growth Rate'] = "N/A"
        result['Next Dividend Date'] = 'N/A'
        result['Predicted Dividend Amount'] = 'N/A'
        result['Dividend Percentage'] = "N/A"
        result['Past Dividends'] = []
    else:
        try:
            result['Dividend Growth Rate'] = dividends.pct_change().mean()
        except Exception:
            result['Dividend Growth Rate'] = "N/A"
            
        result['Latest Close Price'] = latest_close_price
        result['Dividend Percentage'] = "N/A"
        
        try:
            predicted_dividend_amount = dividends.iloc[-1]
            if latest_close_price != "N/A":
                dividend_percentage = (predicted_dividend_amount / latest_close_price) * 100
                result['Dividend Percentage'] = dividend_percentage
            
            past_dividends = dividends.tail(10)
            result['Past Dividends'] = past_dividends.tolist()
            
            date_diffs = past_dividends.index.to_series().diff().dropna()
            if not date_diffs.empty:
                avg_diff = date_diffs.mean()
                last_dividend_date = past_dividends.index[-1]
                next_dividend_date = last_dividend_date + avg_diff
                result['Next Dividend Date'] = str(next_dividend_date.date())
            else:
                result['Next Dividend Date'] = 'N/A'

            result['Predicted Dividend Amount'] = predicted_dividend_amount
        except Exception:
            result['Next Dividend Date'] = 'N/A'
            result['Predicted Dividend Amount'] = 'N/A'
            result['Dividend Percentage'] = "N/A"
            result['Past Dividends'] = []

    # Get actual dividend data from CSV
    try:
        actual_dividends = pd.read_csv(ACTUAL_DIVIDEND_FILE)
        actual_dividends['Symbol'] = actual_dividends['Symbol'].str.upper()
        actual_data = actual_dividends[actual_dividends['Symbol'] == yf_ticker.upper()]
        
        if not actual_data.empty:
            latest_actual = actual_data.iloc[0]
            result['Actual Dividend'] = latest_actual['Dividened Per Share']
            result['Actual Ex Date'] = latest_actual['Ex Date']
            
            # Compare predicted vs actual
            if (result['Predicted Dividend Amount'] != "N/A" and 
                result['Actual Dividend'] != "N/A" and
                float(result['Actual Dividend']) != 0):
                
                prediction_error = abs(float(result['Predicted Dividend Amount']) - float(result['Actual Dividend']))
                result['Dividend Prediction Error'] = prediction_error
                
                # Calculate accuracy percentage (fixed the parenthesis issue here)
                accuracy = 100 - (prediction_error / float(result['Actual Dividend']) * 100)
                result['Prediction Accuracy (%)'] = accuracy
            else:
                result['Dividend Prediction Error'] = "N/A"
                result['Prediction Accuracy (%)'] = "N/A"
        else:
            result['Actual Dividend'] = "N/A"
            result['Actual Ex Date'] = "N/A"
            result['Dividend Prediction Error'] = "N/A"
            result['Prediction Accuracy (%)'] = "N/A"
    except Exception as e:
        st.error(f"Error loading actual dividend data for {ticker}: {e}")
        result['Actual Dividend'] = "N/A"
        result['Actual Ex Date'] = "N/A"
        result['Dividend Prediction Error'] = "N/A"
        result['Prediction Accuracy (%)'] = "N/A"

    return result

# Function to save results to an Excel file
def save_to_excel(results, filename="dividend_predictions.xlsx"):
    try:
        results_df = pd.DataFrame(results)
        
        # Clean up the Past Dividends column for Excel
        if 'Past Dividends' in results_df.columns:
            results_df['Past Dividends'] = results_df['Past Dividends'].apply(
                lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x
            )
        
        if os.path.exists(filename):
            book = load_workbook(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = book
            # Clear existing sheet if it exists
            if 'Dividend Predictions' in book.sheetnames:
                std = book.get_sheet_by_name('Dividend Predictions')
                book.remove_sheet(std)
            results_df.to_excel(writer, index=False, sheet_name='Dividend Predictions')
            writer.save()
        else:
            results_df.to_excel(filename, index=False)
        st.success(f"Results saved to {filename}")
    except Exception as e:
        st.error(f"Error saving to Excel: {e}")

# Streamlit App
st.set_page_config(page_title="Stock Dividend Predictions", layout="wide")

# Display Header Logo
st.markdown("""
    <style>
        .header-logo {
            display: block;
            margin-left: auto;
            margin-right: auto;
            width: 25%;
        }
        /* Hide GitHub icons and fork button */
        .css-1v0mbdj { 
            display: none !important;
        }
        .css-1b22hs3 {
            display: none !important;
        }
        /* Hide Streamlit footer elements */
        footer { 
            display: none !important; 
        }
        /* Hide the GitHub repository button */
        .css-1r6ntm8 { 
            display: none !important;
        }
        /* Style the dataframe */
        .dataframe {
            width: 100%;
        }
        /* Style the progress bar */
        .stProgress > div > div > div > div {
            background-color: #4CAF50;
        }
    </style>
    <img class="header-logo" src="https://pystatiq.com/images/pystatIQ_logo.png" alt="Header Logo">
""", unsafe_allow_html=True)

st.title('Stock Dividend Prediction and Financial Analysis')

# Read the stock symbols from the local stocks.xlsx file
if os.path.exists(STOCKS_FILE_PATH):
    symbols_df = pd.read_excel(STOCKS_FILE_PATH)

    # Check if the 'Symbol' column exists
    if 'Symbol' not in symbols_df.columns:
        st.error("The file must contain a 'Symbol' column with stock tickers.")
    else:
        # Let the user select stocks from the file
        stock_options = symbols_df['Symbol'].tolist()
        selected_stocks = st.multiselect("Select Stock Symbols", stock_options)

        # Button to start the data fetching process
        if st.button('Fetch Financial Data') and selected_stocks:
            all_results = []
            progress_bar = st.progress(0)
            status_text = st.empty()
            total_stocks = len(selected_stocks)
            
            for i, ticker in enumerate(selected_stocks):
                progress = (i + 1) / total_stocks
                progress_bar.progress(progress)
                status_text.text(f"Processing {ticker} ({i+1}/{total_stocks})...")
                
                result = get_financial_data(ticker)
                if result is not None:
                    all_results.append(result)
            
            progress_bar.empty()
            status_text.empty()
            
            if all_results:
                st.subheader("Financial Data Results")
                results_df = pd.DataFrame(all_results)
                
                # Calculate accuracy metrics
                valid_errors = results_df[
                    (results_df['Dividend Prediction Error'] != "N/A") & 
                    (results_df['Actual Dividend'] != "N/A")
                ]
                
                if not valid_errors.empty:
                    st.write(f"### Prediction Accuracy Summary")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Stocks with valid predictions", len(valid_errors))
                    with col2:
                        st.metric("Mean Absolute Error", f"{valid_errors['Dividend Prediction Error'].mean():.2f}")
                    with col3:
                        st.metric("Average Accuracy", f"{valid_errors['Prediction Accuracy (%)'].mean():.2f}%")
                
                # Display the full results
                st.dataframe(results_df)
                
                # Button to save the results to Excel
                if st.button('Save Results to Excel'):
                    save_to_excel(all_results)
                    st.balloons()

else:
    st.error(f"{STOCKS_FILE_PATH} not found. Please ensure the file exists.")

# Display Footer
st.markdown("""
    <div style="text-align: center; font-size: 14px; margin-top: 30px;">
        <p><strong>App Code:</strong> Stock-Dividend-Prediction-Jan-2025</p>
        <p>To get access to the stocks file to upload, please Email us at <a href="mailto:support@pystatiq.com">support@pystatiq.com</a>.</p>
        <p>Don't forget to add the Application code.</p>
        <p><strong>README:</strong> <a href="https://pystatiq-lab.gitbook.io/docs/python-apps/stock-dividend-predictions" target="_blank">https://pystatiq-lab.gitbook.io/docs/python-apps/stock-dividend-predictions</a></p>
    </div>
""", unsafe_allow_html=True)

# Display Footer Logo
st.markdown(f"""
    <style>
        .footer-logo {{
            display: block;
            margin-left: auto;
            margin-right: auto;
            width: 90px;
            padding-top: 30px;
        }}
    </style>
    <img class="footer-logo" src="https://predictram.com/images/logo.png" alt="Footer Logo">
""", unsafe_allow_html=True)
